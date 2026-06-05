"""
Kaizen GO — Calculadora de Desconto Ótimo por Fornecedor
Otimização de portfólio: maximiza MC total sujeito a Fat total >= Fat_atual * (1 + meta_recuperacao).
Lambda calculado como MC_antes / Fat_antes (período de referência).
MC sempre positiva por fornecedor (restrição dura).
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
from scipy.optimize import minimize
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Desconto Ótimo — Kaizen GO", page_icon="🎯", layout="wide")
st.title("🎯 Calculadora de Desconto Ótimo por Fornecedor")
st.caption("Otimização de portfólio — maximiza MC total com meta de recuperação de faturamento.")

# ── Upload do arquivo ─────────────────────────────────────────────────────────
arquivo = st.file_uploader("📂 Selecione o arquivo da análise de fornecedores", type=["csv", "xlsx"])

if arquivo is None:
    st.info("Aguardando upload do arquivo...")
    st.stop()

# ── Leitura do arquivo ────────────────────────────────────────────────────────
try:
    if arquivo.name.endswith('.xlsx'):
        df = pd.read_excel(arquivo, header=0)
    else:
        df = pd.read_csv(arquivo, sep=None, engine='python')
    df.columns = [c.lstrip('\ufeff').strip() for c in df.columns]
except Exception as e:
    st.error(f"Erro ao ler o arquivo: {e}")
    st.stop()

# ── Verificar colunas necessárias ─────────────────────────────────────────────
colunas_necessarias = ['Fornecedor', 'Fat/dia Atual', 'MC/dia Atual', 'Desc Atual %', 'ε Obs']
faltando = [c for c in colunas_necessarias if c not in df.columns]
if faltando:
    st.error(f"Colunas não encontradas: {faltando}")
    st.write("Colunas disponíveis:", df.columns.tolist())
    st.stop()

# ── Converter numéricos ───────────────────────────────────────────────────────
for col in ['Fat/dia Atual', 'MC/dia Atual', 'Desc Atual %', 'ε Obs', 'Fat/dia Antes', 'MC/dia Antes']:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

df = df.dropna(subset=['Fornecedor', 'Fat/dia Atual', 'MC/dia Atual', 'Desc Atual %'])

# ── Sidebar: parâmetros ───────────────────────────────────────────────────────
st.sidebar.title("⚙️ Parâmetros")

meta_recuperacao = st.sidebar.slider(
    "Meta de recuperação de Fat (%)",
    min_value=0, max_value=50, value=10, step=5,
    help="Quanto % acima do Fat/dia atual você quer atingir. Ex: 10% = Fat_alvo = Fat_atual × 1.10"
)

usar_margem_minima = st.sidebar.checkbox("Aplicar margem mínima por fornecedor", value=False)
if usar_margem_minima:
    margem_minima = st.sidebar.number_input(
        "Margem mínima aceitável (%)", min_value=0.0, max_value=30.0, value=5.0, step=0.5,
    )
else:
    margem_minima = 0.0

eps_padrao = st.sidebar.slider(
    "ε padrão (quando não calculado)", min_value=-15.0, max_value=-0.1, value=-2.0, step=0.1,
)
max_pp = st.sidebar.number_input(
    "Máximo de pp a testar", min_value=1.0, max_value=20.0, value=6.0, step=1.0,
)

# ── Funções auxiliares ────────────────────────────────────────────────────────
def fat_nova(fat_d, eps, desc_atual, delta, denom):
    novo_desc = desc_atual - delta
    r = max((1.0 + novo_desc / 100.0) / denom, 0.001)
    return fat_d * (r ** (1.0 + eps))

def mc_nova(fat_d, mc_d, eps, desc_atual, delta, denom):
    custo_d = fat_d - mc_d
    novo_desc = desc_atual - delta
    r = max((1.0 + novo_desc / 100.0) / denom, 0.001)
    f = fat_d * (r ** (1.0 + eps))
    c = custo_d * (r ** eps)
    return f - c

# ── Preparar dados ────────────────────────────────────────────────────────────
df['tem_eps']   = df['ε Obs'].notna()
df['eps_usado'] = df['ε Obs'].where(df['tem_eps'], eps_padrao)
df['denom']     = (1.0 + df['Desc Atual %'] / 100.0).clip(lower=0.001)

df['fat_antes_ref'] = df['Fat/dia Antes'].where(
    df['Fat/dia Antes'].notna() & (df['Fat/dia Antes'] > 0),
    df['Fat/dia Atual']
)

# MC antes — estima pela margem atual aplicada ao fat antes se não tiver
if 'MC/dia Antes' in df.columns and df['MC/dia Antes'].notna().any():
    df['mc_antes_ref'] = df['MC/dia Antes'].where(
        df['MC/dia Antes'].notna() & (df['MC/dia Antes'] > 0),
        df['MC/dia Atual'] / df['Fat/dia Atual'].replace(0, np.nan) * df['fat_antes_ref']
    )
else:
    margem_pct = df['MC/dia Atual'] / df['Fat/dia Atual'].replace(0, np.nan)
    df['mc_antes_ref'] = margem_pct * df['fat_antes_ref']

df['queda_volume'] = (df['Fat/dia Atual'] - df['fat_antes_ref']) / df['fat_antes_ref']

# ── Lambda: MC_antes / Fat_antes ─────────────────────────────────────────────
mc_antes_total  = df['mc_antes_ref'].sum()
fat_antes_total = df['fat_antes_ref'].sum()
lam = mc_antes_total / fat_antes_total if fat_antes_total > 0 else 0.07

# Fat alvo: Fat_atual * (1 + meta_recuperacao%)
fat_atual_total = df['Fat/dia Atual'].sum()
fat_alvo = fat_atual_total * (1 + meta_recuperacao / 100.0)

st.sidebar.divider()
st.sidebar.markdown(f"**λ calculado:** {lam:.4f}")
st.sidebar.markdown(f"**Fat/dia atual:** R$ {fat_atual_total:,.0f}".replace(",","."))
st.sidebar.markdown(f"**Fat/dia alvo (+{meta_recuperacao}%):** R$ {fat_alvo:,.0f}".replace(",","."))
st.sidebar.markdown(f"**Fat/dia Antes (ref):** R$ {fat_antes_total:,.0f}".replace(",","."))
st.sidebar.caption("λ = MC_antes/Fat_antes. Fat alvo = Fat_atual × (1 + meta%).")

# ── Otimização de portfólio ───────────────────────────────────────────────────
rows = df.reset_index(drop=True)
n_forn = len(rows)

# Bounds por fornecedor
bounds = []
for i, row in rows.iterrows():
    queda = row['queda_volume']
    if queda < 0:
        max_subida = max(0.5, max_pp * (1 + queda))
    else:
        max_subida = max_pp
    bounds.append((-max_subida, max_pp))

# Função objetivo: minimiza -(MC_total + λ * Fat_total)
def objetivo(deltas):
    total = 0.0
    for i, row in rows.iterrows():
        if not row['tem_eps']:
            total += row['MC/dia Atual'] + lam * row['Fat/dia Atual']
            continue
        mc  = mc_nova(row['Fat/dia Atual'], row['MC/dia Atual'],
                      row['eps_usado'], row['Desc Atual %'], deltas[i], row['denom'])
        fat = fat_nova(row['Fat/dia Atual'], row['eps_usado'],
                       row['Desc Atual %'], deltas[i], row['denom'])
        total += mc + lam * fat
    return -total

# Restrições
constraints = []

# 1. Fat total >= fat_alvo
def restricao_fat(deltas):
    total_fat = 0.0
    for i, row in rows.iterrows():
        if not row['tem_eps']:
            total_fat += row['Fat/dia Atual']
            continue
        total_fat += fat_nova(row['Fat/dia Atual'], row['eps_usado'],
                              row['Desc Atual %'], deltas[i], row['denom'])
    return total_fat - fat_alvo

constraints.append({'type': 'ineq', 'fun': restricao_fat})

# 2. MC_i >= margem_minima% do fat (e sempre positiva)
for i, row in rows.iterrows():
    if row['tem_eps']:
        def mc_minima(deltas, idx=i, r=row):
            fat = fat_nova(r['Fat/dia Atual'], r['eps_usado'],
                          r['Desc Atual %'], deltas[idx], r['denom'])
            mc  = mc_nova(r['Fat/dia Atual'], r['MC/dia Atual'],
                         r['eps_usado'], r['Desc Atual %'], deltas[idx], r['denom'])
            return mc - max(margem_minima / 100.0, 0.001) * fat
        constraints.append({'type': 'ineq', 'fun': mc_minima})

# Ponto inicial inteligente — fornecedores que caíram muito começam com desconto inicial
x0 = np.array([
    min(max_pp * 0.3, bounds[i][1]) if rows.iloc[i]['queda_volume'] < -0.10 else 0.0
    for i in range(n_forn)
])

with st.spinner("Otimizando portfólio..."):
    resultado = minimize(
        objetivo, x0,
        method='SLSQP',
        bounds=bounds,
        constraints=constraints,
        options={'ftol': 1e-6, 'maxiter': 3000}
    )

deltas_otimos = resultado.x

# ── Montar resultados ─────────────────────────────────────────────────────────
resultados = []
for i, row in rows.iterrows():
    fat_d      = row['Fat/dia Atual']
    mc_d       = row['MC/dia Atual']
    desc_atual = row['Desc Atual %']
    eps        = row['eps_usado']
    tem_eps    = row['tem_eps']
    denom      = row['denom']

    if tem_eps:
        delta_otimo = deltas_otimos[i]
        fat_otimo   = fat_nova(fat_d, eps, desc_atual, delta_otimo, denom)
        mc_otima    = mc_nova(fat_d, mc_d, eps, desc_atual, delta_otimo, denom)
    else:
        delta_otimo = 0.0
        fat_otimo   = fat_d
        mc_otima    = mc_d

    desc_otimo   = desc_atual - delta_otimo
    margem_atual = mc_d / fat_d * 100 if fat_d > 0 else 0
    margem_otima = mc_otima / fat_otimo * 100 if fat_otimo > 0 else 0
    ganho_mc     = mc_otima - mc_d

    if abs(delta_otimo) < 0.05:
        acao = "manter"
    elif delta_otimo > 0:
        acao = "↑ mais desconto"
    else:
        acao = "↓ menos desconto"

    resultados.append({
        'Fornecedor':        row['Fornecedor'],
        'Fat/dia Antes':     row.get('Fat/dia Antes', None),
        'Fat/dia Atual':     fat_d,
        'MC/dia Atual':      mc_d,
        'Margem Atual %':    margem_atual,
        'Desc Atual %':      desc_atual,
        'ε Obs':             row['ε Obs'] if tem_eps else None,
        'Tem ε':             tem_eps,
        'Queda Volume %':    round(row['queda_volume'] * 100, 1),
        'Δ Desc Ótimo (pp)': round(delta_otimo, 1),
        'Ação':              acao,
        'Desc Ótimo %':      round(desc_otimo, 1),
        'Fat/dia Ótimo':     fat_otimo,
        'MC/dia Ótima':      mc_otima,
        'Margem Ótima %':    margem_otima,
        'Ganho MC/dia':      ganho_mc,
    })

res = pd.DataFrame(resultados)

# ── KPIs gerais ───────────────────────────────────────────────────────────────
st.divider()
st.markdown("### 📊 Resumo Geral")

GANHO_MIN = 1.0

mc_atual_total  = res['MC/dia Atual'].sum()
mc_otima_total  = res['MC/dia Ótima'].sum()
ganho_total     = mc_otima_total - mc_atual_total
fat_otimo_total = res['Fat/dia Ótimo'].sum()
n_subir_preco   = int((res['Δ Desc Ótimo (pp)'] < -0.05).sum())
n_mais_desc     = int((res['Δ Desc Ótimo (pp)'] > 0.05).sum())

status_fat = "✅ Meta atingida" if fat_otimo_total >= fat_alvo * 0.99 else "⚠️ Abaixo da meta"

k1, k2, k3, k4, k5, k6 = st.columns(6)
k1.metric("MC/dia Atual",                f"R$ {mc_atual_total:,.0f}".replace(",","."))
k2.metric("MC/dia Ótima",                f"R$ {mc_otima_total:,.0f}".replace(",","."),
          f"+R$ {ganho_total:,.0f}/dia".replace(",","."))
k3.metric("Fat/dia Ótimo",               f"R$ {fat_otimo_total:,.0f}".replace(",","."), status_fat)
k4.metric(f"Fat/dia Alvo (+{meta_recuperacao}%)", f"R$ {fat_alvo:,.0f}".replace(",","."))
k5.metric("Sugerem ↓ desconto",          f"{n_subir_preco}")
k6.metric("Sugerem ↑ desconto",          f"{n_mais_desc}")

if not resultado.success:
    gap = fat_alvo - fat_otimo_total
    st.warning(
        f"⚠️ Otimizador não convergiu completamente. "
        f"Meta: R$ {fat_alvo:,.0f} · Atingido: R$ {fat_otimo_total:,.0f} · "
        f"Gap: R$ {gap:,.0f} ({gap/fat_alvo*100:.1f}%). "
        f"Tente reduzir a meta de recuperação."
    )
else:
    st.success("✅ Otimização convergiu!")

# ── Tabela de resultados ──────────────────────────────────────────────────────
st.divider()
st.markdown("### 📋 Desconto Ótimo por Fornecedor")
st.caption(
    f"λ = {lam:.4f} (MC_antes/Fat_antes) · ε padrão: {eps_padrao} · "
    f"Max pp: {max_pp:.0f} · Fat alvo: R$ {fat_alvo:,.0f} (+{meta_recuperacao}% do atual)".replace(",",".")
)

disp = res[[
    'Fornecedor', 'Fat/dia Antes', 'Fat/dia Atual', 'MC/dia Atual', 'Margem Atual %',
    'Desc Atual %', 'ε Obs', 'Queda Volume %',
    'Δ Desc Ótimo (pp)', 'Ação', 'Desc Ótimo %',
    'Fat/dia Ótimo', 'MC/dia Ótima', 'Margem Ótima %', 'Ganho MC/dia'
]].copy().sort_values('Ganho MC/dia', ascending=False)

def cor_ganho(v):
    try:
        fv = float(v)
        if fv > GANHO_MIN:  return 'background-color:#c6efce; color:#0b5e1e; font-weight:bold'
        if fv < -GANHO_MIN: return 'background-color:#ffd6d6; color:#6b0000'
    except: pass
    return ''

def cor_acao(v):
    if v == '↓ menos desconto': return 'color:#006600; font-weight:bold'
    if v == '↑ mais desconto':  return 'color:#003399; font-weight:bold'
    return 'color:#888888'

fmt = {
    'Fat/dia Antes':     lambda v: f"R$ {v:,.0f}".replace(",",".") if pd.notna(v) else "—",
    'Fat/dia Atual':     lambda v: f"R$ {v:,.0f}".replace(",",".") if pd.notna(v) else "—",
    'MC/dia Atual':      lambda v: f"R$ {v:,.0f}".replace(",",".") if pd.notna(v) else "—",
    'Margem Atual %':    lambda v: f"{v:.1f}%" if pd.notna(v) else "—",
    'Desc Atual %':      lambda v: f"{v:.1f}%" if pd.notna(v) else "—",
    'ε Obs':             lambda v: f"{v:.2f}" if pd.notna(v) else "—",
    'Queda Volume %':    lambda v: f"{v:+.1f}%" if pd.notna(v) else "—",
    'Δ Desc Ótimo (pp)': lambda v: f"{v:+.1f} pp" if pd.notna(v) else "—",
    'Desc Ótimo %':      lambda v: f"{v:.1f}%" if pd.notna(v) else "—",
    'Fat/dia Ótimo':     lambda v: f"R$ {v:,.0f}".replace(",",".") if pd.notna(v) else "—",
    'MC/dia Ótima':      lambda v: f"R$ {v:,.0f}".replace(",",".") if pd.notna(v) else "—",
    'Margem Ótima %':    lambda v: f"{v:.1f}%" if pd.notna(v) else "—",
    'Ganho MC/dia':      lambda v: f"R$ {v:+,.0f}".replace(",",".") if pd.notna(v) else "—",
}

st.dataframe(
    disp.style.format(fmt, na_rep="—")
        .map(cor_ganho, subset=['Ganho MC/dia'])
        .map(cor_acao,  subset=['Ação']),
    use_container_width=True,
    height=min(600, 60 + len(disp) * 35)
)

# ── Exportar Excel ────────────────────────────────────────────────────────────
st.divider()
st.markdown("### ⬇️ Exportar Resultado")

def gerar_excel(df_res):
    wb = Workbook()
    ws = wb.active
    ws.title = "Desconto Ótimo"
    cols = list(df_res.columns)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=9)
        cell.fill = PatternFill('solid', start_color='2F4F8F')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 40

    for ri, row in enumerate(df_res.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            v = round(float(val), 2) if isinstance(val, float) and not pd.isna(val) else (val if pd.notna(val) else None)
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(horizontal='center')
            col_name = cols[ci - 1]
            if col_name == 'Ganho MC/dia' and isinstance(val, float) and not pd.isna(val):
                cor = 'C6EFCE' if val > GANHO_MIN else ('FFD6D6' if val < -GANHO_MIN else 'FFFFFF')
                cell.fill = PatternFill('solid', start_color=cor)

    ws.column_dimensions['A'].width = 20
    for ci in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

excel_buf = gerar_excel(res)
st.download_button(
    "⬇️ Baixar Excel com desconto ótimo",
    excel_buf,
    "desconto_otimo_fornecedores.xlsx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)